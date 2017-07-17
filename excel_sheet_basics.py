import openpyxl
import os

os.chdir('C:\\Users\\aryan\\PycharmProjects\\classes')
wb=openpyxl.load_workbook('example.xlsx')
sheet=wb.get_sheet_names()
print (type(sheet),' IS TYPE OF SHEET ', wb, ' IS WB OJECT ', sheet , ' IS SHEET')
sheet=wb.get_sheet_by_name('Sheet3')

print ('title of sheet', sheet.title)
active_sheet=wb.active
print ('ACTIVE SHEEET IS:-',active_sheet, 'its type', type(active_sheet))

print ('ROW ',active_sheet['A1'],  ' HAS ', active_sheet['A1'].value, ' AS VALUE AND \
ITS CO-ORDINATE ARE:-', active_sheet['A1'].coordinate)

print ('************** THE ROW , COLUMN AND ITS VALUES ***********')
maxrow=active_sheet.max_row
maxcolumn=active_sheet.max_column

print (active_sheet,'HAS', maxrow, ' MAX NO OF ROWS AND', maxcolumn, 'MAX COLUMNS' )

for rowi in range(1,maxrow+1):
    for columnj in range(1,maxcolumn+1):
        print ('ROW:-',rowi, 'COLUMN:-', columnj ,'ITS VALUE ARE:-',active_sheet.cell(row=rowi, column=columnj).value)
    print ('************ END OF ROW *******************')