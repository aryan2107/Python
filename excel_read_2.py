import os
import ntpath as path
import openpyxl

path=r'C:\Users\aryan\PycharmProjects\classes'

os.chdir(path)

wb=openpyxl.load_workbook('censuspopdata.xlsx')
active=wb.active
print ('ACTIVE SHEET IS:-', active)

maxr=active.max_row
maxc=active.max_column
print ('MAXROW AND MAXCOLUMN OF THE SHEET ARE:- ',maxr, maxc)
maxr=10
countyData={}
countydata,state,censustract,county, pop2010={}, {}, {}, {}, {}
for j in range(1,maxr):


     print('STATE:-', active['B' + str(j)].value)
     state['A' + str(j)] = active['B' + str(j)].value

     print('COUNTY:-', active['C' + str(j)].value)
     county['A' + str(j)] = active['C' + str(j)].value
     print('POP2010:-', active['D' + str(j)].value)
     pop2010['A' + str(j)] = active['D' + str(j)].value
     print('CENSUSTRACT:-', active['A' + str(j)].value)
     censustract['A' + str(j)] = active['A' + str(j)].value
     print('********END OF ROW*******************')

     #print ('COLUMN NO:-',j, ' is ', active.cell(row=1,column=j).value)
pop=0

for i in range(2,maxr+1):
    pop+=int(active.cell(row=i, column=4).value)

print ('TOTAL POPULATION IS:-',pop)

print (censustract, end='\n')
print (state, end='\n')
print (county, end='\n')
print (pop2010, end='\n')
print (countyData, end='\n')
cenToState={}
l=[]
for i in state:
    print ('i ' , i)
    last=state[i]
    if state[i] == last:
        print (state[i],' is state[i]', i, ' iteration', ' and last is', last)
        print (censustract[i], ' is censustract[i]', i, ' iteration')
        cenToState[i]=l.append(censustract[i])
        print ('l is:-',l)
        last=state[i]
    else:
        l=[]

print (cenToState)





